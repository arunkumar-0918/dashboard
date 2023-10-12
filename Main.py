import streamlit as st
import pandas as pd
import folium
from streamlit_folium import st_folium
from folium  import plugins
import numpy as np
import openpyxl



st.set_page_config(page_title="CROC-Contracts_overview",page_icon="😎",layout="wide")
st.title("CROC Contracts Dashboard")
wb_obj = openpyxl.load_workbook("Geocoodrinates of all countries.xlsx")
sheet =wb_obj.active

# Extract data from rows, skipping the first row (for example, skip 1 row)
data = []
for row in sheet.iter_rows(values_only=True):
    data.append(row)

# Create a Pandas DataFrame
geo_coordinates  = pd.DataFrame(data[1:],columns=data[0])
#geo_coordinates = pd.DataFrame(wb_obj.active.values)

wb_obj1 = openpyxl.load_workbook("June.xlsx")

sheet =wb_obj1.active

# Extract data from rows, skipping the first row (for example, skip 1 row)
data = []
for row in sheet.iter_rows(values_only=True):
    data.append(row)

# Create a Pandas DataFrame
MCL  = pd.DataFrame(data[1:],columns=data[0])
#MCL = pd.DataFrame(wb_obj1.active.values)

Final_dashboard=pd.merge(geo_coordinates,MCL,left_on="name",right_on="Region",indicator="True")
totalcounts=len(Final_dashboard)
st.header(f"Total Contracts in CROC: {totalcounts}")
col1, col2, col3 = st.columns(3)
st.sidebar.header("Filter here based on contract type, Market & on/offboarded")
contract_type=st.sidebar.multiselect(
    "Select the contract type:",
    options=Final_dashboard["Contract Type"].unique(),
    default=Final_dashboard["Contract Type"].unique()
)

Contract_Status=st.sidebar.multiselect(
    "Select the Contract_Status:",
    options=Final_dashboard["Status"].unique(),
    default=Final_dashboard["Status"].unique()
)




MCL_selection=Final_dashboard[Final_dashboard['Contract Type'].isin(contract_type) & Final_dashboard['Status'].isin(Contract_Status)]


with col1:
      Total_dynamcially_counts= len(MCL_selection)
      st.write(f"Market_wise_Contracts: {Total_dynamcially_counts}")

#onnectitcut_center= (41.5025,-72.699997)
#map = folium.Map(location=[MCL_selection['latitude'].mean(), MCL_selection['longitude'].mean()], Zoom=2.5)
map = folium.Map(location=[20.593684, 78.96288], Zoom=100)
for  index,row in MCL_selection.iterrows():
      #location_coordinates=MCL_selection['latitude'],MCL_selection['longitude']
      folium.Marker([row['latitude'],row['longitude']], Zoom=2.5).add_to(map)


plugins.Fullscreen(position='topright').add_to(map)

st_folium(map,width=1500)